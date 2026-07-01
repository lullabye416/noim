# -*- coding: utf-8 -*-
"""
xls_convert.py — .xls → .xlsx 변환 (환경 편차 방어)

문제
----
기존 gui.py는 .xls 변환에 반드시 MS Excel(win32com)이 필요했다.
Excel 미설치/버전 상이/gen_py 캐시 손상 PC에서 즉시 실패한다.

해결
----
1순위: Excel COM (레이아웃 100% 보존, 가장 안전)
2순위: xlrd + openpyxl 순수 파이썬 변환 (Excel 없어도 값은 옮김)
       ※ 서식 일부 손실 가능 → WARN 로그로 명시

두 경로 모두 실패하면 원인을 로그로 남기고 예외를 올린다.
"""

import os

try:
    from app_logging import get_logger, capture_exception
    _log = get_logger("xls_convert")
except Exception:
    import logging
    _log = logging.getLogger("xls_convert")
    def capture_exception(exc=None, **ctx):
        _log.error("예외 %s", ctx, exc_info=exc or True)


def convert_xls_to_xlsx(src_path: str) -> str:
    """
    src_path(.xls) → 같은 위치의 _conv.xlsx 반환.
    반환된 경로는 임시 파일이므로 호출측에서 정리 대상에 넣을 것.
    """
    abs_in  = os.path.abspath(src_path)
    abs_out = os.path.splitext(abs_in)[0] + "_conv.xlsx"

    # ── 1순위: Excel COM ──────────────────────────────────────────
    try:
        return _convert_via_com(abs_in, abs_out)
    except Exception as e:
        _log.warning("Excel COM 변환 실패 → 순수 파이썬 폴백 시도: %s: %s",
                     type(e).__name__, e)

    # ── 2순위: xlrd 폴백 ──────────────────────────────────────────
    try:
        return _convert_via_xlrd(abs_in, abs_out)
    except Exception as e:
        capture_exception(e, step="xls_convert_fallback", file=os.path.basename(src_path))
        raise RuntimeError(
            f".xls 변환 실패. Excel이 설치되지 않았고 자동 변환도 실패했습니다.\n"
            f"원본을 Excel에서 직접 .xlsx로 저장 후 다시 시도하세요.\n원인: {e}"
        )


def _convert_via_com(abs_in: str, abs_out: str) -> str:
    import win32com.client as win32
    import pythoncom
    excel = wb = None
    try:
        pythoncom.CoInitialize()
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(abs_in)
        wb.SaveAs(abs_out, FileFormat=51)   # 51 = xlsx
        _log.info("Excel COM 변환 완료 → %s", os.path.basename(abs_out))
        return abs_out
    finally:
        if wb:    wb.Close(SaveChanges=False)
        if excel: excel.Quit()
        pythoncom.CoUninitialize()


def _convert_via_xlrd(abs_in: str, abs_out: str) -> str:
    """
    Excel 없이 값만 옮기는 폴백. 서식/수식은 보존되지 않음.
    노임일보 입력은 '값'만 필요하므로 대부분 문제없다.
    """
    import xlrd
    import openpyxl

    book = xlrd.open_workbook(abs_in)
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    for sheet in book.sheets():
        ws = wb_out.create_sheet(title=sheet.name[:31])  # 시트명 31자 제한
        for r in range(sheet.nrows):
            for c in range(sheet.ncols):
                ws.cell(row=r + 1, column=c + 1, value=sheet.cell_value(r, c))

    wb_out.save(abs_out)
    _log.warning("xlrd 폴백 변환 완료(서식 손실 가능) → %s", os.path.basename(abs_out))
    return abs_out