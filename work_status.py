# -*- coding: utf-8 -*-
"""
작업현황(직영) 시트 자동 채우기 로직
"""
import re
import calendar
import datetime
from copy import copy

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.styles import PatternFill

try:
    from app_logging import get_logger
    _log = get_logger("work_status")
except Exception:
    import logging
    _log = logging.getLogger("work_status")

_NO_FILL = PatternFill(fill_type=None)

# ── 상수 ─────────────────────────────────────────────────────────
WINTER_MONTHS     = {10, 11, 1, 2}
STATUS_DATA_START = 4       # 1일 데이터 시작 행 (1-based)
NOIM_DATA_START   = 8       # 노임일보 데이터 시작 행
NOIM_NAME_COL     = 2       # 노임일보 이름 열 (B)
NOIM_TODAY_COL    = 4       # 노임일보 금일 열 (D) — 0이면 미출근
NOIM_WORKDESC_COL = 9       # 노임일보 작업내용 열 (I)

# 샘플 기준 열 위치 (동절기 있는 상태)
COL_TOTAL   = 3   # C 전체
COL_JIK     = 4   # D 직영
COL_SHIN    = 5   # E 신호수
COL_ANJEON  = 6   # F 안전
COL_SELYUN  = 7   # G 세륜기
COL_DONG    = 8   # H 동절기
COL_DESC    = 9   # I 작업사항

# 직종 분류: 카테고리 키 → 기본 포함 키워드/표시값 (탭2 UI "4. 작업내용 매핑 규칙"에서 재정의 가능)
DEFAULT_CATEGORY_KEYWORDS = {'shin': '신호수', 'anjeon': '안전자재', 'selyun': '세륜기', 'dong': '동절기'}
DEFAULT_CATEGORY_VALUES = {
    'shin': '3번게이트 신호수', 'anjeon': '안전자재 정리', 'selyun': '세륜기 관리', 'dong': '동절기',
}
DEFAULT_FALLBACK_VALUE = '현장정리'
DESC_ORDER = ['shin', 'anjeon', 'selyun', 'dong']

# 세륜기/계륜기는 현장에서 혼용되는 동의어라, 사용자가 세륜기 키워드를 바꿔도 항상 함께 매칭한다.
_SELYUN_SYNONYM = '계륜기'


def build_category_rules(category_keywords: dict = None) -> list:
    """{category: keyword} → [(keyword, category), ...] (분류 우선순위 순서)."""
    ck = category_keywords if category_keywords else DEFAULT_CATEGORY_KEYWORDS
    rules = []
    for cat in ('shin', 'anjeon'):
        kw = (ck.get(cat) or '').strip()
        if kw:
            rules.append((kw, cat))
    selyun_kw = (ck.get('selyun') or '').strip()
    if selyun_kw:
        rules.append((selyun_kw, 'selyun'))
    rules.append((_SELYUN_SYNONYM, 'selyun'))
    dong_kw = (ck.get('dong') or '').strip()
    if dong_kw:
        rules.append((dong_kw, 'dong'))
    return rules


# ── 노임일보 파싱 ─────────────────────────────────────────────────

def _classify(work_desc: str, category_rules: list) -> str:
    for kw, cat in category_rules:
        if kw in work_desc:
            return cat
    return 'jik'


def read_noim_ilbo(noim_file: str, category_rules: list = None) -> dict:
    """
    {day: {'total','shin','anjeon','selyun','dong','jik'}}
    """
    category_rules = category_rules if category_rules is not None else build_category_rules()
    wb = openpyxl.load_workbook(noim_file, data_only=True)
    day_data = {}

    for sheet_name in wb.sheetnames:
        if not re.fullmatch(r'\d{2}', sheet_name):
            continue
        day = int(sheet_name)
        ws  = wb[sheet_name]

        counts = {'shin': 0, 'anjeon': 0, 'selyun': 0, 'dong': 0, 'jik': 0}
        total  = 0

        for r in range(NOIM_DATA_START, NOIM_DATA_START + 60):
            name_val = ws.cell(r, NOIM_NAME_COL).value
            if not name_val or not str(name_val).strip():
                break                                    # 이름 없으면 데이터 끝
            today_val = ws.cell(r, NOIM_TODAY_COL).value
            try:
                worked = today_val is not None and float(today_val) > 0
            except (ValueError, TypeError):
                worked = False
            if not worked:
                continue                                 # 금일=0 → 미출근, 카운트 제외
            gongsu = float(today_val)
            total += gongsu
            work_desc = str(ws.cell(r, NOIM_WORKDESC_COL).value or '').strip()
            counts[_classify(work_desc, category_rules)] += gongsu

        if total > 0:
            day_data[day] = {**counts, 'total': total}

    wb.close()
    return day_data


# ── 유틸 ─────────────────────────────────────────────────────────

def _build_desc_line(info: dict, is_winter: bool, category_values: dict, fallback_value: str) -> str:
    """공수가 잡힌 카테고리만, 현재 매핑 표의 '값'을 그대로 이어붙인다.
    (노임일보 파일에 적힌 원문이 아니라 실행 시점의 매핑 값을 쓰므로,
    탭1을 언제 돌렸는지와 무관하게 탭2 실행 시점 기준 값이 반영된다.)"""
    order = DESC_ORDER if is_winter else [k for k in DESC_ORDER if k != 'dong']
    result = []
    for cat in order:
        if info.get(cat, 0) > 0:
            val = (category_values.get(cat) or '').strip()
            if val and val not in result:
                result.append(val)
    if info.get('jik', 0) > 0:
        val = (fallback_value or '').strip()
        if val and val not in result:
            result.append(val)
    return ', '.join(result)


def _copy_row_style(ws, src_row: int, dst_row: int, max_col: int = 10):
    for col in range(1, max_col + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        if src.has_style:
            dst.font          = copy(src.font)
            dst.border        = copy(src.border)
            dst.fill          = copy(src.fill)
            dst.number_format = src.number_format
            dst.protection    = copy(src.protection)
            dst.alignment     = copy(src.alignment)


def _scan_sheet(ws):
    """col1(NO) 기준 day→row 매핑과 합계행 위치 반환."""
    day_to_row = {}
    sum_row    = None
    for r in range(STATUS_DATA_START, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v is None:
            continue
        if isinstance(v, (int, float)) and float(v) == int(float(v)) and int(float(v)) > 0:
            day_to_row[int(float(v))] = r
        elif str(v).strip() == '계':
            sum_row = r
            break
    return day_to_row, sum_row


def _delete_col_with_widths(ws, col_idx: int):
    """
    열 삭제 후 column_dimensions(열 너비)를 수동으로 한 칸씩 당김.
    openpyxl이 너비 정보를 자동으로 보정하지 않을 때를 대비.
    """
    # 삭제 전 너비 저장
    saved = {}
    for letter, dim in ws.column_dimensions.items():
        saved[column_index_from_string(letter)] = dim.width

    ws.delete_cols(col_idx, 1)

    # 삭제된 열 이후 너비를 한 칸씩 앞으로 당김
    for idx in sorted(saved.keys()):
        if idx < col_idx:
            ws.column_dimensions[get_column_letter(idx)].width = saved[idx]
        elif idx > col_idx:
            ws.column_dimensions[get_column_letter(idx - 1)].width = saved[idx]


# ── 메인 함수 ─────────────────────────────────────────────────────

def fill_work_status(noim_file: str, output_file: str, sheet_name: str,
                     year: int, month: int, category_keywords: dict = None,
                     category_values: dict = None, fallback_value: str = None) -> str:
    """
    output_file: 샘플을 복사한 수정 대상 파일 (호출 전 shutil.copy 필요).
    category_keywords/category_values/fallback_value: 탭2 "4. 작업내용 매핑 규칙" UI에서
        전달되는 {카테고리: 키워드}/{카테고리: 표시값}/기타행 값. 키워드는 분류(공수 카운트)에,
        표시값은 작업사항 열 문구 렌더링에 쓰인다 — 후자는 노임일보 원문이 아니라 실행 시점의
        매핑 표 값을 그대로 쓰므로 탭1 재실행 여부와 무관하게 항상 최신 값이 반영된다.
        미지정 시 DEFAULT_CATEGORY_KEYWORDS/DEFAULT_CATEGORY_VALUES/DEFAULT_FALLBACK_VALUE 사용.
    """
    category_rules = build_category_rules(category_keywords)
    display_values = category_values if category_values else DEFAULT_CATEGORY_VALUES
    fallback_value = fallback_value if fallback_value else DEFAULT_FALLBACK_VALUE
    day_data    = read_noim_ilbo(noim_file, category_rules)
    is_winter   = month in WINTER_MONTHS
    needed_days = calendar.monthrange(year, month)[1]

    wb = openpyxl.load_workbook(output_file)
    ws = wb[sheet_name]

    # ── 1. 시트 구조 파악 ─────────────────────────────────────────
    day_to_row, sum_row = _scan_sheet(ws)
    current_days = max(day_to_row.keys()) if day_to_row else 0

    # ── 2. 행 수 조정 ─────────────────────────────────────────────
    if needed_days < current_days:
        last_needed_row = day_to_row[needed_days]
        del_count = (sum_row - 1) - last_needed_row
        if del_count > 0:
            ws.delete_rows(last_needed_row + 1, del_count)
            sum_row -= del_count

    elif needed_days > current_days:
        last_row  = day_to_row[current_days]
        add_count = needed_days - current_days
        ws.insert_rows(last_row + 1, add_count)
        for i in range(add_count):
            new_row = last_row + 1 + i
            _copy_row_style(ws, last_row, new_row)
            ws.cell(new_row, 1).value = current_days + 1 + i
        if sum_row:
            sum_row += add_count

    # ── 3. 날짜 채우기 ────────────────────────────────────────────
    for day_no in range(1, needed_days + 1):
        ws.cell(STATUS_DATA_START + day_no - 1, 2).value = datetime.datetime(year, month, day_no)

    # ── 4. 데이터 쓰기 (동절기 포함 원래 열 번호 기준) ───────────
    for day_no in range(1, needed_days + 1):
        r    = STATUS_DATA_START + day_no - 1
        info = day_data.get(day_no)
        if info:
            ws.cell(r, COL_TOTAL ).value = info['total']
            ws.cell(r, COL_JIK   ).value = info['jik']
            ws.cell(r, COL_SHIN  ).value = info['shin']
            ws.cell(r, COL_ANJEON).value = info['anjeon']
            ws.cell(r, COL_SELYUN).value = info['selyun']
            ws.cell(r, COL_DONG  ).value = info['dong']
            ws.cell(r, COL_DESC  ).value = _build_desc_line(info, is_winter, display_values, fallback_value) or None
        else:
            for c in [COL_TOTAL, COL_JIK, COL_SHIN, COL_ANJEON, COL_SELYUN, COL_DONG]:
                ws.cell(r, c).value = 0
            ws.cell(r, COL_DESC).value = None

    # ── 5. 합계 행 수식 (동절기 열 포함 상태에서 먼저 작성) ──────
    if sum_row:
        first = STATUS_DATA_START
        last  = STATUS_DATA_START + needed_days - 1
        for cl in ['C', 'D', 'E', 'F', 'G', 'H']:
            ws[f'{cl}{sum_row}'] = f'=SUM({cl}{first}:{cl}{last})'

    # ── 6. 합계 행 숨김 해제 ─────────────────────────────────────
    if sum_row:
        ws.row_dimensions[sum_row].hidden = False

    # ── 7. 비동절기: 동절기 열 삭제 + 합계 수식 H 제거 ──────────
    if not is_winter:
        _delete_col_with_widths(ws, COL_DONG)
        # 동절기 삭제 후 새 H열(작업사항)은 텍스트 → 합계행 H셀 비움
        if sum_row:
            ws.cell(sum_row, 8).value = None

    # ── 8. 여분 열 노란 fill 제거 ────────────────────────────────
    # 비고(col9 비동절기 / col10 동절기) 이후 열은 템플릿 잔여 스타일 제거
    max_content_col = 9 if not is_winter else 10
    for r in range(STATUS_DATA_START, STATUS_DATA_START + needed_days + 2):
        for c in range(max_content_col + 1, ws.max_column + 1):
            ws.cell(r, c).fill = _NO_FILL

    # ── 9. 작성일 쓰기 ────────────────────────────────────────────
    for c in range(1, ws.max_column + 1):
        v = ws.cell(2, c).value
        if v and '작성일' in str(v):
            ws.cell(2, c).value = f' 작성일 :  {month}월 {needed_days}일'
            break

    wb.save(output_file)
    wb.close()
    _log.info("작업현황 채우기 완료 | 시트=%s | %d-%02d | 데이터일수=%d | 동절기=%s",
              sheet_name, year, month, len(day_data), is_winter)
    if not day_data:
        _log.warning("노임일보에서 읽은 출근 데이터가 없습니다 — 입력 파일/시트명을 확인하세요.")
    return output_file